import io
from datetime import datetime
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from selecition_engine import (
    MASTER_ALIASES, PARTY_ALIASES, DEFAULT_RATIO_STANDARDS,
    guess_column, guess_header_row, guess_shape_mapping, norm_str,
    build_requirement_df, build_bucket_index, prepare_party_df, allocate,
    split_results,
)

st.set_page_config(page_title="Diamond Auto-Selection", page_icon="💎", layout="wide")

# ----------------------------------------------------------------------
# Session state defaults
# ----------------------------------------------------------------------
defaults = {
    'req_df': None,            # live requirement grid (Need Remaining decreases over time)
    'req_df_original': None,   # untouched snapshot, for reset
    'master_shapes': [],
    'ratio_standards': DEFAULT_RATIO_STANDARDS.copy(),
    'run_done': False,
    'selection_df': None,
    'rejection_df': None,
    'party_filename': None,
    'history': [],             # list of dicts: filename, selected, rejected, timestamp
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


def df_to_excel_bytes(sheets: dict) -> bytes:
    """sheets: {sheet_name: dataframe}. Returns styled xlsx bytes."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.sheets[name[:31]]
            header_fill = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
            header_font = Font(name='Arial', bold=True, color='FFFFFF')
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).head(200)])
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 45)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.font = Font(name='Arial')
            ws.freeze_panes = 'A2'
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions
    return buf.getvalue()


st.title("💎 Diamond Auto-Selection")
st.caption(
    "Master requirement grid → party-file matching → Selection / Rejection / Mix export. "
    "Designed to handle party files that come in any column layout."
)

# ======================================================================
# SIDEBAR — Ratio standards (editable, global)
# ======================================================================
with st.sidebar:
    st.header("⚙️ Settings")
    st.subheader("Standard ratio bands")
    st.caption("Used only when a party file's ratio check is set to **Standard**. Add a row per shape as needed.")
    st.session_state['ratio_standards'] = st.data_editor(
        st.session_state['ratio_standards'],
        num_rows="dynamic",
        width='stretch',
        key="ratio_std_editor",
    )
    undefined_shape_allow = st.radio(
        "If a shape has no ratio band defined above, in Standard mode:",
        ["Allow it through", "Reject it"],
        index=0,
        key="undefined_shape_policy",
    ) == "Allow it through"

    st.divider()
    if st.session_state['req_df'] is not None:
        st.metric("Total Need Remaining", int(st.session_state['req_df']['Need Remaining'].sum()))
        if st.button("🔄 Reset requirement grid to original Grid/Available"):
            st.session_state['req_df'] = st.session_state['req_df_original'].copy()
            st.session_state['run_done'] = False
            st.rerun()

# ======================================================================
# STEP 1 — Master file
# ======================================================================
st.header("Step 1 — Upload Master File (requirement grid)")
master_file = st.file_uploader("Master / company report file (.xlsx)", type=["xlsx", "xls"], key="master_uploader")

if master_file is not None:
    master_bytes = master_file.getvalue()
    try:
        xls = pd.ExcelFile(io.BytesIO(master_bytes))
    except Exception as e:
        st.error(f"Could not open file: {e}")
        st.stop()
    sheet = xls.sheet_names[0]
    if len(xls.sheet_names) > 1:
        sheet = st.selectbox("Sheet", xls.sheet_names, key="master_sheet")

    raw_preview = pd.read_excel(io.BytesIO(master_bytes), sheet_name=sheet, header=None, nrows=15)
    guessed_hdr = guess_header_row(raw_preview)
    with st.expander("Preview raw file & choose header row", expanded=False):
        st.dataframe(raw_preview, width='stretch')
    hdr_row = st.number_input(
        "Header row (1 = first row of the sheet)", min_value=1, max_value=15,
        value=guessed_hdr + 1, key="master_hdr_row",
    )
    master_df = pd.read_excel(io.BytesIO(master_bytes), sheet_name=sheet, header=hdr_row - 1)
    master_df = master_df.dropna(axis=1, how='all').dropna(how='all')

    st.markdown("**Map master-file columns**")
    cols = list(master_df.columns)
    c1, c2, c3, c4 = st.columns(4)
    c5, c6, c7 = st.columns(3)
    field_widgets = {}
    layout = [c1, c2, c3, c4, c5, c6, c7]
    field_labels = {
        'shape': 'Shape', 'from_size': 'From Size', 'to_size': 'To Size',
        'clarity': 'Clarity', 'color': 'Color', 'grid': 'Grid (max pcs needed)',
        'available': 'Available (in stock)',
    }
    for i, (field, label) in enumerate(field_labels.items()):
        guess = guess_column(cols, MASTER_ALIASES[field])
        options = ['-- select --'] + cols
        idx = options.index(guess) if guess in options else 0
        with layout[i]:
            field_widgets[field] = st.selectbox(label, options, index=idx, key=f"master_map_{field}")

    missing = [f for f, v in field_widgets.items() if v == '-- select --']
    if missing:
        st.warning(f"Please map: {', '.join(field_labels[m] for m in missing)}")
    else:
        if st.button("✅ Build requirement grid", type="primary"):
            colmap = {f: v for f, v in field_widgets.items()}
            req_df = build_requirement_df(master_df, colmap)
            st.session_state['req_df'] = req_df
            st.session_state['req_df_original'] = req_df.copy()
            st.session_state['master_shapes'] = sorted(req_df['Shape'].unique().tolist())
            st.session_state['run_done'] = False
            st.success(f"Requirement grid built — {len(req_df)} buckets, "
                       f"total Need to Buy = {int(req_df['Need to Buy'].sum())} pcs.")

if st.session_state['req_df'] is not None:
    req_df = st.session_state['req_df']
    with st.expander("📊 Requirement grid (live — Need Remaining updates as you process party files)", expanded=False):
        c1, c2, c3 = st.columns(3)
        c1.metric("Buckets", len(req_df))
        c2.metric("Total Grid (target)", int(req_df['Grid'].sum()))
        c3.metric("Total Need Remaining", int(req_df['Need Remaining'].sum()))
        by_shape = req_df.groupby('Shape')[['Grid', 'Available', 'Need to Buy', 'Need Remaining']].sum()
        st.bar_chart(by_shape['Need Remaining'])
        st.dataframe(
            req_df[['Shape', 'Size Group', 'Clarity', 'Color', 'Grid', 'Available',
                    'Need to Buy', 'Need Remaining']],
            width='stretch', height=300,
        )

# ======================================================================
# STEP 2 — Party file
# ======================================================================
if st.session_state['req_df'] is not None:
    st.header("Step 2 — Upload Party File (stones offered for sale)")
    party_file = st.file_uploader(
        "Party file — any format/column layout (.xlsx)", type=["xlsx", "xls"], key="party_uploader",
    )

    if party_file is not None:
        party_bytes = party_file.getvalue()
        try:
            pxls = pd.ExcelFile(io.BytesIO(party_bytes))
        except Exception as e:
            st.error(f"Could not open file: {e}")
            st.stop()
        psheet = pxls.sheet_names[0]
        if len(pxls.sheet_names) > 1:
            psheet = st.selectbox("Sheet", pxls.sheet_names, key="party_sheet")

        praw_preview = pd.read_excel(io.BytesIO(party_bytes), sheet_name=psheet, header=None, nrows=20)
        guessed_phdr = guess_header_row(praw_preview)
        with st.expander("Preview raw party file & choose header row", expanded=True):
            st.dataframe(praw_preview, width='stretch', height=250)
        phdr_row = st.number_input(
            "Header row (1 = first row of the sheet)", min_value=1, max_value=20,
            value=guessed_phdr + 1, key="party_hdr_row",
        )
        party_df = pd.read_excel(io.BytesIO(party_bytes), sheet_name=psheet, header=phdr_row - 1)
        party_df = party_df.dropna(axis=1, how='all').dropna(how='all')
        st.caption(f"{len(party_df)} stone rows detected, {len(party_df.columns)} usable columns.")

        pcols = list(party_df.columns)
        st.markdown("**Map required columns**")
        c1, c2, c3, c4 = st.columns(4)
        req_fields = {'shape': 'Shape *', 'color': 'Color *', 'clarity': 'Clarity *', 'carat': 'Carat / Size *'}
        pmap = {}
        for col_box, (field, label) in zip([c1, c2, c3, c4], req_fields.items()):
            guess = guess_column(pcols, PARTY_ALIASES[field])
            options = ['-- select --'] + pcols
            idx = options.index(guess) if guess in options else 0
            pmap[field] = col_box.selectbox(label, options, index=idx, key=f"party_map_{field}")

        st.markdown("**Cut / Polish / Symmetry**")
        cc1, cc2, cc3, cc4 = st.columns(4)
        for col_box, field in zip([cc1, cc2, cc3], ['cut', 'polish', 'symmetry']):
            guess = guess_column(pcols, PARTY_ALIASES[field])
            options = ['-- none in this file --'] + pcols
            idx = options.index(guess) if guess in options else 0
            pmap[field] = col_box.selectbox(field.capitalize(), options, index=idx, key=f"party_map_{field}")
        cps_mode = cc4.radio(
            "Cut/Polish/Symmetry rule", ["Standard", "Specific"], key="cps_mode",
            help="Standard = only ID/Ideal or EX/Excellent pass (blank also passes). "
                 "Specific = take any grade.",
        )
        for f in ['cut', 'polish', 'symmetry']:
            if pmap[f] == '-- none in this file --':
                pmap[f] = None

        st.markdown("**Ratio**")
        rc1, rc2, rc3 = st.columns(3)
        ratio_source_label = rc1.selectbox(
            "Ratio source",
            ["Existing ratio column", "Calculate from Length & Width columns",
             "Calculate from Measurement text (L x W x H)", "No ratio data / skip"],
            key="ratio_source_label",
        )
        ratio_source_map = {
            "Existing ratio column": 'existing',
            "Calculate from Length & Width columns": 'length_width',
            "Calculate from Measurement text (L x W x H)": 'measurement',
            "No ratio data / skip": 'none',
        }
        ratio_source = ratio_source_map[ratio_source_label]
        if ratio_source == 'existing':
            guess = guess_column(pcols, PARTY_ALIASES['ratio'])
            options = ['-- select --'] + pcols
            pmap['ratio'] = rc2.selectbox("Ratio column", options,
                                           index=options.index(guess) if guess in options else 0,
                                           key="party_map_ratio")
        elif ratio_source == 'length_width':
            gL = guess_column(pcols, PARTY_ALIASES['length'])
            gW = guess_column(pcols, PARTY_ALIASES['width'])
            optionsL = ['-- select --'] + pcols
            pmap['length'] = rc2.selectbox("Length column", optionsL,
                                            index=optionsL.index(gL) if gL in optionsL else 0,
                                            key="party_map_length")
            pmap['width'] = rc3.selectbox("Width column", optionsL,
                                           index=optionsL.index(gW) if gW in optionsL else 0,
                                           key="party_map_width")
        elif ratio_source == 'measurement':
            gM = guess_column(pcols, PARTY_ALIASES['measurement'])
            optionsM = ['-- select --'] + pcols
            pmap['measurement'] = rc2.selectbox("Measurement column", optionsM,
                                                 index=optionsM.index(gM) if gM in optionsM else 0,
                                                 key="party_map_measurement")
        ratio_check_mode = rc3.radio(
            "Ratio rule", ["Standard", "Specific"], key="ratio_check_mode",
            help="Standard = must fall inside the band set in the sidebar for that shape. "
                 "Specific = any ratio accepted.",
        ) if ratio_source != 'none' else 'Specific'

        st.markdown("**Allocation priority (optional)**")
        sc1, sc2 = st.columns(2)
        sort_col = sc1.selectbox("Prefer to fill requirement using (optional sort)",
                                  ['-- file order --'] + pcols, key="sort_col")
        sort_asc = sc2.radio("Order", ["Ascending", "Descending"], key="sort_asc") == "Ascending"
        sort_col = None if sort_col == '-- file order --' else sort_col

        required_missing = [v for v in [pmap['shape'], pmap['color'], pmap['clarity'], pmap['carat']]
                             if v == '-- select --']
        if required_missing:
            st.warning("Please map Shape, Color, Clarity and Carat/Size before continuing.")
        else:
            st.markdown("**Step 3 — Map party shapes → requirement-grid shapes**")
            distinct_shapes = party_df[pmap['shape']].dropna().unique().tolist()
            master_shapes = st.session_state['master_shapes']
            shape_map_rows = []
            for s in distinct_shapes:
                guess = guess_shape_mapping(s, master_shapes)
                shape_map_rows.append({'Party Shape Value': str(s), 'Maps To': guess or 'SKIP / Not in grid'})
            shape_map_df = pd.DataFrame(shape_map_rows)
            shape_map_df = st.data_editor(
                shape_map_df,
                column_config={
                    'Maps To': st.column_config.SelectboxColumn(
                        options=['SKIP / Not in grid'] + master_shapes,
                    )
                },
                disabled=['Party Shape Value'],
                width='stretch',
                key="shape_map_editor",
                hide_index=True,
            )
            shape_map = {
                norm_str(r['Party Shape Value']): (None if r['Maps To'] == 'SKIP / Not in grid' else r['Maps To'])
                for _, r in shape_map_df.iterrows()
            }

            if st.button("🚀 Run auto-selection", type="primary"):
                bucket_index = build_bucket_index(st.session_state['req_df'])
                mapping = dict(pmap)
                mapping['_bucket_index'] = bucket_index
                modes = {'cps_mode': cps_mode, 'ratio_source': ratio_source, 'ratio_check_mode': ratio_check_mode}
                prepped = prepare_party_df(
                    party_df, mapping, shape_map, modes, st.session_state['ratio_standards'],
                    undefined_shape_allow, master_shapes,
                )
                result_df, new_req = allocate(prepped, st.session_state['req_df'], sort_col=sort_col, ascending=sort_asc)
                selection_df, rejection_df = split_results(result_df, list(party_df.columns), new_req)

                st.session_state['req_df'] = new_req
                st.session_state['selection_df'] = selection_df
                st.session_state['rejection_df'] = rejection_df
                st.session_state['run_done'] = True
                st.session_state['party_filename'] = party_file.name
                st.session_state['history'].append({
                    'file': party_file.name,
                    'selected': len(selection_df),
                    'rejected': len(rejection_df),
                    'time': datetime.now().strftime('%Y-%m-%d %H:%M'),
                })
                st.rerun()

# ======================================================================
# STEP 3 — Results / downloads
# ======================================================================
if st.session_state['run_done']:
    st.header("Step 3 — Results")
    sel, rej = st.session_state['selection_df'], st.session_state['rejection_df']
    m1, m2, m3 = st.columns(3)
    m1.metric("Selected", len(sel))
    m2.metric("Rejected", len(rej))
    m3.metric("Need Remaining (after this file)", int(st.session_state['req_df']['Need Remaining'].sum()))

    tab1, tab2, tab3 = st.tabs(["✅ Selection", "❌ Rejection", "📊 Requirement summary"])
    with tab1:
        st.dataframe(sel, width='stretch', height=350)
    with tab2:
        st.dataframe(rej, width='stretch', height=350)
        st.caption("Reject reason breakdown")
        st.dataframe(rej['Reason'].value_counts().rename_axis('Reason').reset_index(name='Count'),
                     width='stretch')
    with tab3:
        st.dataframe(
            st.session_state['req_df'][['Shape', 'Size Group', 'Clarity', 'Color', 'Grid', 'Available',
                                         'Need to Buy', 'Need Remaining']],
            width='stretch', height=350,
        )

    st.subheader("⬇️ Downloads")
    base_name = (st.session_state['party_filename'] or 'party_file').rsplit('.', 1)[0]
    summary_df = st.session_state['req_df'][['Shape', 'Size Group', 'Clarity', 'Color', 'Grid',
                                              'Available', 'Need to Buy', 'Need Remaining']]

    d1, d2, d3 = st.columns(3)
    d1.download_button(
        "Selection.xlsx", df_to_excel_bytes({'Selection': sel}),
        file_name=f"Selection_{base_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    d2.download_button(
        "Rejection.xlsx", df_to_excel_bytes({'Rejection': rej}),
        file_name=f"Rejection_{base_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    d3.download_button(
        "Mix.xlsx (Selection + Rejection + Summary)",
        df_to_excel_bytes({'Selection': sel, 'Rejection': rej, 'Requirement Summary': summary_df}),
        file_name=f"Mix_{base_name}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if st.session_state['history']:
        with st.expander("Processing history this session"):
            st.dataframe(pd.DataFrame(st.session_state['history']), width='stretch')
